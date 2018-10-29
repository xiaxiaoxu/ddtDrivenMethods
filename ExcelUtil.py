#encoding=utf-8
from openpyxl import load_workbook

class ParseExcel(object):

    def __init__(self, excelPath, sheetName):
        # 将要读取的excel加载到内存
        self.wb = load_workbook(excelPath)
        # 通过工作表名称获取一个工作表对象
        self.sheet = self.wb.get_sheet_by_name(sheetName)
        # 获取工作表中存在数据的区域的最大行号
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        # 用于存放从工作表中读取出来的数据
        dataList = []
        # 因为工作表中的第一行是标题行，所以需要去掉
        for line in self.sheet.rows:  
            # 遍历工作表中数据区域的每一行，
            # 并将每行中各个单元格的数据取出存于列表tmpList中，
            # 然后再将存放一行数据的列表添加到最终数据列表dataList中
            tmpList = []
            tmpList.append(line[1].value)
            tmpList.append(line[2].value)
            dataList.append(tmpList)
        # 将获取工作表中的所有数据的迭代对象返回
        return dataList[1:]

if __name__ == '__main__':
    excelPath = u'E:\\数据驱动\\测试数据.xlsx'
    sheetName = u"搜索数据表"
    pe = ParseExcel(excelPath, sheetName)
    print pe.getDatasFromSheet()
    for i in pe.getDatasFromSheet():
        print i[0], i[1]
